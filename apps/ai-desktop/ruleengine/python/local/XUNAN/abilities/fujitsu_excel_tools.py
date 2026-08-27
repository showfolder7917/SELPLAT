#!/usr/bin/env python3
"""Fujitsu DB、SQL 规格书和基本设计 Excel 生成能力。

统一替代五个 Java 入口，使用 openpyxl、标准 XML/JSON/CSV 解析完成离线处理。
所有命令都要求显式输入输出，默认配置只从当前用户规则包定位。
"""

from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
import json
from pathlib import Path
import re
import shutil
from typing import Iterable
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = next(path for path in Path(__file__).resolve().parents if (path / "settings.gradle").is_file())
FUJITSU_RESOURCE_ROOT = (
    PROJECT_ROOT
    / "apps/ai-desktop/ruleengine/rules/local/XUNAN/fujitsu"
)
SQL_TEMPLATE_ROOT = (
    FUJITSU_RESOURCE_ROOT
    / "通用/template/RUL_FujitsuSQL规格书Excel生成规则/SQL仕様書生成ツール"
)
DEFAULT_SQL_CONFIG = SQL_TEMPLATE_ROOT / "SQL仕様書生成ツール新規.json"


@dataclass(frozen=True)
class SqlStatement:
    """保存一条 Mapper SQL 的稳定审查字段。"""

    statement_id: str
    operation: str
    sql: str
    source: str


def locate_default_config() -> Path:
    """返回当前用户 Fujitsu 规则包中的唯一默认配置。"""

    if not DEFAULT_SQL_CONFIG.is_file():
        raise FileNotFoundError(f"默认 SQL 规格书配置不存在：{DEFAULT_SQL_CONFIG}")
    return DEFAULT_SQL_CONFIG


def _load_config(path: Path | None) -> dict[str, object]:
    target = path or locate_default_config()
    return json.loads(target.read_text(encoding="utf-8"))


def extract_mapper_sql(source_paths: Iterable[Path]) -> list[SqlStatement]:
    """从 MyBatis XML 和 Java 注解中提取 SQL，不执行数据库连接。"""

    statements: list[SqlStatement] = []
    for source in source_paths:
        if source.suffix.lower() == ".xml":
            statements.extend(_extract_xml_mapper(source))
        elif source.suffix.lower() == ".java":
            statements.extend(_extract_java_mapper(source))
    return statements


def _extract_xml_mapper(path: Path) -> list[SqlStatement]:
    """解析 select/insert/update/delete 元素并保留动态标签文字。"""

    root = ET.parse(path).getroot()
    result: list[SqlStatement] = []
    for element in root.iter():
        operation = element.tag.rsplit("}", 1)[-1]
        if operation not in {"select", "insert", "update", "delete"}:
            continue
        sql = " ".join("".join(element.itertext()).split())
        result.append(SqlStatement(str(element.attrib.get("id") or ""), operation, sql, str(path)))
    return result


def _extract_java_mapper(path: Path) -> list[SqlStatement]:
    """识别 MyBatis 四类注解中的字符串 SQL。"""

    text = path.read_text(encoding="utf-8")
    pattern = re.compile(r"@(Select|Insert|Update|Delete)\s*\(\s*(?:\{\s*)?\"(.*?)\"", re.DOTALL)
    result: list[SqlStatement] = []
    for index, match in enumerate(pattern.finditer(text), 1):
        sql = bytes(match.group(2), "utf-8").decode("unicode_escape") if "\\" in match.group(2) else match.group(2)
        result.append(SqlStatement(f"annotation_{index}", match.group(1).lower(), " ".join(sql.split()), str(path)))
    return result


def generate_sql_spec(
    config_path: Path | None,
    source_paths: list[Path],
    output_path: Path,
    *,
    template_path: Path | None = None,
) -> dict[str, object]:
    """按模板生成 SQL 规格书；没有模板时创建同结构新工作簿。"""

    config = _load_config(config_path)
    statements = extract_mapper_sql(source_paths)
    configured_template = config.get("templatePath") or config.get("template_path")
    template = template_path or (Path(str(configured_template)) if configured_template else None)
    if template and not template.is_absolute():
        template = PROJECT_ROOT / template
    if template and template.is_file():
        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(template, output_path)
        workbook = load_workbook(output_path, keep_vba=output_path.suffix.lower() == ".xlsm")
    else:
        from openpyxl import Workbook
        workbook = Workbook()
    sheet = workbook[workbook.sheetnames[0]]
    sheet.title = str(config.get("sheetName") or "SQL仕様書")[:31]
    headers = ["ID", "操作", "SQL", "来源"]
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(1, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row, statement in enumerate(statements, 2):
        values = [statement.statement_id, statement.operation, statement.sql, statement.source]
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row, column, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 100
    sheet.column_dimensions["D"].width = 55
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return {"statements": len(statements), "output": str(output_path)}


def correct_sql_spec(source_path: Path, output_path: Path) -> dict[str, object]:
    """修正规格书的换行、边框、筛选和冻结窗格，保留既有单元格值。"""

    workbook = load_workbook(source_path, keep_vba=source_path.suffix.lower() == ".xlsm")
    thin = Side(style="thin", color="B8C2CC")
    changed = 0
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        if sheet.max_row >= 1 and sheet.max_column >= 1:
            sheet.auto_filter.ref = sheet.dimensions
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical="top",
                    wrap_text=True,
                )
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                changed += 1
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return {"changed_cells": changed, "output": str(output_path)}


def import_database(workbook_path: Path, output_root: Path) -> dict[str, object]:
    """把每个非空工作表导出为 UTF-8 CSV，作为可审查的数据库测试数据。"""

    workbook = load_workbook(workbook_path, data_only=False, read_only=True)
    output_root.mkdir(parents=True, exist_ok=True)
    files: list[str] = []
    for sheet in workbook.worksheets:
        target = output_root / f"{_safe_name(sheet.title)}.csv"
        with target.open("w", encoding="utf-8-sig", newline="") as file_obj:
            writer = csv.writer(file_obj)
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(["" if value is None else value for value in row])
        files.append(str(target))
    return {"sheets": len(files), "files": files}


def export_database(source_root: Path, workbook_path: Path) -> dict[str, object]:
    """把目录内 CSV 合并成一个工作簿，每个文件对应一个工作表。"""

    from openpyxl import Workbook
    csv_files = sorted(source_root.glob("*.csv"))
    if not csv_files:
        raise FileNotFoundError(f"没有可导出的 CSV：{source_root}")
    workbook = Workbook()
    workbook.remove(workbook.active)
    for csv_path in csv_files:
        sheet = workbook.create_sheet(_safe_name(csv_path.stem)[:31])
        with csv_path.open("r", encoding="utf-8-sig", newline="") as file_obj:
            for row in csv.reader(file_obj):
                sheet.append(row)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(workbook_path)
    return {"sheets": len(csv_files), "output": str(workbook_path)}


def generate_design_document(kind: str, input_json: Path, output_path: Path) -> dict[str, object]:
    """把 API 概览或接口定义 JSON 生成为统一基本设计 Excel。"""

    from openpyxl import Workbook
    payload = json.loads(input_json.read_text(encoding="utf-8"))
    records = payload if isinstance(payload, list) else payload.get("items") or payload.get("apis") or [payload]
    keys: list[str] = []
    for record in records:
        if isinstance(record, dict):
            for key in record:
                if key not in keys:
                    keys.append(key)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "API概要" if kind == "api-overview" else "インターフェース仕様"
    sheet.append(keys)
    for record in records:
        sheet.append([json.dumps(record.get(key), ensure_ascii=False) if isinstance(record.get(key), (dict, list)) else record.get(key) for key in keys])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = 28
        for cell in column:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return {"records": len(records), "output": str(output_path)}


def _safe_name(value: str) -> str:
    return re.sub(r"[\\/*?:\[\]]", "_", value).strip() or "Sheet"


def main() -> int:
    """提供五类旧 Java 程序的统一 Python 子命令。"""

    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="action", required=True)
    sql = subparsers.add_parser("sql-generate")
    sql.add_argument("output", type=Path)
    sql.add_argument("sources", type=Path, nargs="+")
    sql.add_argument("--config", type=Path)
    sql.add_argument("--template", type=Path)
    correct = subparsers.add_parser("sql-correct")
    correct.add_argument("source", type=Path)
    correct.add_argument("output", type=Path)
    db_import = subparsers.add_parser("db-import")
    db_import.add_argument("workbook", type=Path)
    db_import.add_argument("output_root", type=Path)
    db_export = subparsers.add_parser("db-export")
    db_export.add_argument("source_root", type=Path)
    db_export.add_argument("workbook", type=Path)
    for name in ("api-overview", "interface-spec"):
        design = subparsers.add_parser(name)
        design.add_argument("input_json", type=Path)
        design.add_argument("output", type=Path)
    arguments = parser.parse_args()
    if arguments.action == "sql-generate":
        result = generate_sql_spec(arguments.config, arguments.sources, arguments.output, template_path=arguments.template)
    elif arguments.action == "sql-correct":
        result = correct_sql_spec(arguments.source, arguments.output)
    elif arguments.action == "db-import":
        result = import_database(arguments.workbook, arguments.output_root)
    elif arguments.action == "db-export":
        result = export_database(arguments.source_root, arguments.workbook)
    else:
        result = generate_design_document(arguments.action, arguments.input_json, arguments.output)
    print(json.dumps({"status": "completed", **result}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
